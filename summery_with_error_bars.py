from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series 
from openpyxl.utils import get_column_letter
from openpyxl.chart.marker import Marker
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.data_source import StrRef
from openpyxl.drawing.colors import ColorChoice

from openpyxl.chart.error_bar import ErrorBars
from openpyxl.chart.data_source import NumDataSource
from openpyxl.chart.data_source import NumRef
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties

def  Headers():
        headers = [
        "Sheet Name", "Epsilon", "Q", "G",
        "Avg - Width", "Std - Width", "Std/Avg - Width",
        "Avg - Length", "Std - Length", "Std/Avg - Length",
        "Avg - Theta", "Std - Theta", "Std/Avg - Theta",
        "Avg - L/W", "Std - L/W", "Std/Avg - L/W",
        "Error value - Width", "Error value - Length", "Error value - STD/AVG Width","Error value - STD/AVG Length","Error value - AVG L/W", "Error value - STD/AVG L/W"
    ]
        return headers

def process_excel():
    """
    Reads an Excel file sheet by sheet and creates a new file with metrics.
    Returns the number of sheets processed.
    """
    # Check if the summary sheet already exists and remove it if it does
    if summary_sheet_name in input_wb.sheetnames:
        del input_wb[summary_sheet_name]
    
    # Create a new sheet for the summary
    summary_ws = input_wb.create_sheet(title=summary_sheet_name)

    headers = Headers()
    summary_ws.append(headers)

    sheet_count = 0
    for sheet_name in input_wb.sheetnames[:-1]:
        sheet = input_wb[sheet_name]
        epsilon, q, g = None, None, None
        parts = sheet_name.split("_")
        for i, part in enumerate(parts):
            if part == "Epsilon" and i + 1 < len(parts):
                epsilon = float(parts[i + 1])
            elif part == "Q" and i + 1 < len(parts):
                q = float(parts[i + 1])
            elif part == "G" and i + 1 < len(parts):
                g = float(parts[i + 1])

        avg_width = sheet["G2"].value
        std_width = sheet["G3"].value
        ratio_width = std_width / avg_width if avg_width else None
        err_width = Compute_error_cell_size(avg_width, ratio_width, 0.2)
        last_used_row=find_final_row_number(sheet, "B")
        err_std_width= Compute_error_STD(last_used_row, std_width)
        err_std_over_avg_width=Compute_error_STD_over_AVG(err_std_width, std_width, err_width, avg_width, ratio_width)
        print(f"sheet name: {sheet_name} err_std_width: {err_std_width} std: {std_width}")
        

        avg_length = sheet["Y2"].value
        std_length = sheet["Y3"].value
        ratio_length = std_length / avg_length if avg_length else None
        err_length = Compute_error_cell_size(avg_length, ratio_length, 0.2)
        last_used_row=find_final_row_number(sheet, "T")
        err_std_length= Compute_error_STD(last_used_row, std_length)
        err_std_over_avg_length=Compute_error_STD_over_AVG(err_std_length, std_length, err_length, avg_length, ratio_length)

        avg_theta = sheet["AP2"].value
        std_theta = sheet["AP3"].value
        ratio_theta = std_theta / avg_theta if avg_theta else None

        avg_lw = sheet["BJ2"].value
        std_lw = sheet["BJ3"].value
        ratio_lw = std_lw / avg_lw if avg_lw else None
        err_lw=Compute_error_STD_over_AVG(err_length, avg_length, err_width, avg_width, avg_lw)
        last_used_row=find_final_row_number(sheet, "BE")
        err_std_lw= Compute_error_STD(last_used_row, std_lw)
        err_std_over_avg_lw=Compute_error_STD_over_AVG(err_std_lw, std_lw, err_lw, avg_lw, ratio_lw)


    

        summary_ws.append([
            sheet_name, epsilon, q, g,
            avg_width, std_width, ratio_width,
            avg_length, std_length, ratio_length,
            avg_theta, std_theta, ratio_theta,
            avg_lw, std_lw, ratio_lw,
            err_width, err_length, err_std_over_avg_width, err_std_over_avg_length, err_lw, err_std_over_avg_lw
        ])
        sheet_count += 1
    input_wb.save(input_path_to_save)
    print(f"Summary added to {input_path_to_save}")
    return sheet_count , input_wb

def Compute_error_cell_size(avg_value, ratio_value, threshold, high_err_pct=0.15, low_err_pct=0.08):
    if ratio_value is None or avg_value is None:
        err = 0  # or skip this point entirely, or set to some default
        return err
    
    pct = high_err_pct if ratio_value > threshold else low_err_pct

    err = avg_value * pct
    
    return err

def Compute_error_STD(number_of_cells, STD):
    delta_std=STD/((2*(number_of_cells-1))**0.5)
    return delta_std

def Compute_error_STD_over_AVG(delta_std, STD, Delta_avg, avg, ratio):
    error_value=ratio * (((delta_std/STD)**2 + (Delta_avg/avg)**2)**0.5)
    return error_value

def find_final_row_number(sheet, column_letter):
    max_row = sheet.max_row  # maximum number of rows in the sheet

    for row in range(max_row, 0, -1):  # start from bottom
        cell_value = sheet[f"{column_letter}{row}"].value
        if cell_value is not None:
            last_used_row = row
            break

    return last_used_row

def Add_error_bars(ws, copied_section_start, max_row, col):
    if col == 5: #avg width
        error_col = 17
    elif col == 8: #avg length
        error_col = 18
    elif col == 7: #avg std/width
        error_col = 19
    elif col == 10: #avg std/length
        error_col = 20
    elif col == 14: #avg lw
        error_col = 21
    elif col == 16: #avg std/lw
        error_col = 22
    else:
        return None
    error_values = Reference(ws, min_col=error_col, min_row=copied_section_start+2, max_row=max_row)
    error_values_data = NumDataSource(NumRef(error_values))
    error_bars = ErrorBars(errDir='y', errValType='cust', plus=error_values_data, minus=error_values_data)
    error_bars.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="FF0000"))  # Set line color to red
    return error_bars

def create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_column, y_column):
    """
    Creates scatter charts in the same sheet where data is read from.
    Args:
        ws: The worksheet object.
        copied_section_start: The starting row for the copied data.
        coulmn_counter: The column index for placing the chart.
        x_column: The column index for the x-axis data.
        y_column: The column index for the y-axis data.

    """

    # Create a ScatterChart object
    scatter_chart = ScatterChart()
    scatter_chart.title = f"{ws.cell(copied_section_start+1,y_column).value} in corellation to {ws.cell(copied_section_start+1,x_column).value}"
    scatter_chart.style = 13
    scatter_chart.x_axis.title = f"{ws.cell(copied_section_start+1,x_column).value}"
    scatter_chart.y_axis.title = f"{ws.cell(copied_section_start+1,y_column).value}"
    scatter_chart.width = 20
    scatter_chart.height = 15

    # Remove the legend
    scatter_chart.legend = None

    # Determine the range for the data
    max_row = ws.max_row
    x_values = Reference(ws, min_col=x_column, min_row=copied_section_start+2, max_row=max_row)
    y_values = Reference(ws, min_col=y_column, min_row=copied_section_start+2, max_row=max_row)

    # Add the series to the scatter chart
    series = Series(y_values, x_values)
    series.marker = Marker('circle')  # Set marker style to dots
    series.marker.graphicalProperties.solidFill = "FF0000"  # Set marker color to red
    series.graphicalProperties.line.noFill = True  # Disable lines connecting data points

    error_bars=Add_error_bars(ws, copied_section_start, max_row, y_column)
    series.errBars = error_bars



    scatter_chart.series.append(series)

    # Determine chart position to the right of the data

    chart_position = f"{get_column_letter(coulmn_counter)}{copied_section_start}"  # Place chart 11 columns to the right
    ws.add_chart(scatter_chart, chart_position)
    coulmn_counter+=11
    #print(f"Scatter chart added to '{ws.title}'")
    return coulmn_counter

def create_CFD_chart(coulmn_counter, name, copied_section_start=1, filtered_data_row=1):
    """
    Creates scatter plots for each sheet based on the data in columns C (x-axis) and F (y-axis),
    after filtering data with the copy_filtered_data functions.
    """
    ws_input = input_wb[summary_sheet_name]  
    
    scatter_chart = ScatterChart()
    scatter_chart.title = f"cfd for changing {name}"
    scatter_chart.style = 13
    scatter_chart.x_axis.title = "Bins"
    scatter_chart.y_axis.title = "Probability"
    scatter_chart.width = 20
    scatter_chart.height = 15

    colors = [
    'FF0000',  # Red
    '00FF00',  # Green
    '0000FF',  # Blue
    'FFFF00',  # Yellow
    'FF00FF',  # Magenta
    '00FFFF',  # Cyan
    '800080',  # Purple
    '800000',  # Maroon
    '808000',  # Olive
    'A52A2A',  # Brown
    '808080',  # Gray
    'FF6347',  # Tomato
    'FFD700',  # Gold
    'ADFF2F',  # Green Yellow
    'F0E68C',  # Khaki
    'D2691E',  # Chocolate
    '008000',  # Dark Green
    'DC143C',  # Crimson
    '8A2BE2',  # Blue Violet
    'FF1493'   # Deep Pink
]


    # Sort sheets by their order in the workbook
    filterd_sheet_list_names=[]
    for i in range(copied_section_start+2 ,filtered_data_row):
        filterd_sheet_list_names.append(ws_input.cell(i,1).value)

    #print(filterd_sheet_list_names)

    for idx, sheet_name in enumerate(filterd_sheet_list_names):
        if sheet_name in input_wb.sheetnames:
            sheet = input_wb[sheet_name]
        else:
            print(f"Warning: Sheet '{sheet_name}' not found in the workbook!")


        # Check if the sheet has the necessary columns for creating scatter plots (C and F)
        if sheet.max_row > 1:  # Ensure there's data
            x_values = [sheet.cell(row=row, column=3).value for row in range(2, sheet.max_row + 1)]
            y_values = [sheet.cell(row=row, column=6).value for row in range(2, sheet.max_row + 1)]

            if x_values and y_values:  # Ensure there is data in both columns
                # Create scatter chart for each sheet

                # Prepare the x and y values as a Reference
                x_values_ref = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
                y_values_ref = Reference(sheet, min_col=6, min_row=2, max_row=sheet.max_row)

                # Create a Series for the scatter plot
                series = Series(y_values_ref, x_values_ref)
                #print(f"Creating series for sheet: {sheet_name}")
                series.title = SeriesLabel(strRef=StrRef(sheet_name))
                # Set the color for this series (using the colors list, cycling through it)
                color = colors[idx % len(colors)]  # This ensures colors repeat if there are more sheets than colors
                series.graphicalProperties.line.solidFill = color
                scatter_chart.series.append(series)

    # Determine chart position to the right of the data
    chart_position = f"{get_column_letter(coulmn_counter)}{copied_section_start}"
    ws_input.add_chart(scatter_chart, chart_position)
    coulmn_counter += 11  # Move chart position for the next one

    print(f"cfd plot created")
    return coulmn_counter

def copy_and_filter_acordding_to_one_paramter(E_value, Q_value, G_value, sheet_count):
    """
    Copies and filters the data according to E Q G,
    adding them sequentially in rows below the existing data.
    
    Args:
        file_path (str): Path to the output Excel file to modify.
        E_value (str): The E value to filter.
        Q_value (str): The Q value to filter.
        G_value (str): The G value to filter.
        sheet_count (int): Number of sheets in the input file.
    """
    ws = input_wb[summary_sheet_name]
    last_row = ws.max_row
    headers = Headers()

    global first_go  # Tell Python to use the global `first_go`

    if first_go:
        copied_section_start = last_row + 2
    else:
        copied_section_start = last_row + 27
    
    if E_value is None:
        name = "Epsilon"
        value_number_1 = Q_value
        row_number_1 = 2
        value_number_2 = G_value
        row_number_2 = 3
        x_colomn = 2
    elif Q_value is None:
        name = "Q"
        value_number_1 = E_value
        row_number_1 = 1
        value_number_2 = G_value
        row_number_2 = 3
        x_colomn = 3
    elif G_value is None:
        name = "Gamma"
        value_number_1 = E_value
        row_number_1 = 1
        value_number_2 = Q_value
        row_number_2 = 2
        x_colomn = 4
    
    # Create the title and headers for the filtered data
    ws.cell(row=copied_section_start, column=1, value=f"Filtered Data for changing {name}")
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=copied_section_start + 1, column=col_index, value=header)

    filtered_data = []

    # Collect filtered data
    for row in ws.iter_rows(min_row=2, max_row=sheet_count, values_only=True):
        if float(row[row_number_1]) == float(value_number_1) and float(row[row_number_2]) == float(value_number_2):
            filtered_data.append(row)
    
    # Sort the filtered data by the parameter that is not being checked
    if E_value is None:
        filtered_data.sort(key=lambda x: x[x_colomn-1])  # Sort by Q (row_number_1 points to Q)
    elif Q_value is None:
        filtered_data.sort(key=lambda x: x[x_colomn-1])  # Sort by G (row_number_2 points to G)
    elif G_value is None:
        filtered_data.sort(key=lambda x: x[x_colomn-1])  # Sort by E (row_number_1 points to E)

    # Insert the sorted data into the sheet
    filtered_data_row = copied_section_start + 2
    for row in filtered_data:
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=filtered_data_row, column=col_index, value=value)
        filtered_data_row += 1

    # Continue with the scatter charts
    max_column = 18
    coulmn_counter = max_column + 3
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 5)     # avg width
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 7)     # std/avg width
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 8)     # avg length
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 10)    # std/avg length
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 11)    # avg theta
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 12)    # std/avg theta
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 14)    # avg L/W
    coulmn_counter = create_scatter_charts_in_same_sheet(ws, copied_section_start, coulmn_counter, x_colomn, 16)    # std/avg L/W
    coulmn_counter = create_CFD_chart(coulmn_counter, name, copied_section_start, filtered_data_row)

    first_go = False
    input_wb.save(input_path_to_save)
    print(f"Filtered Data for changing {name}")

# Example Usage
input_path_to_save="data_analasys.xlsx"
input_wb = load_workbook(input_path_to_save)
summary_sheet_name="Summery"
sheet_count , input_wb = process_excel()
first_go=True
#copy_filtered_data_Q_G( Q_value="50", G_value="1.2", sheet_count=sheet_count)
#copy_filtered_data_Q_E( Q_value="25", E_value="3.116", sheet_count=sheet_count)
#copy_filtered_data_G_E( G_value="1.2", E_value="3.116", sheet_count=sheet_count)

copy_and_filter_acordding_to_one_paramter(E_value=None,     Q_value="50",   G_value="1.2",   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value=None,     Q_value="50",   G_value="1.3",   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value=None,     Q_value="50",   G_value="1.4",   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value=None,     Q_value="50",   G_value="1.5",   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value=None,     Q_value="25",   G_value="1.4",   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value="3.116",  Q_value=None ,  G_value="1.2",   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value="3.116",  Q_value="25",   G_value=None ,   sheet_count=sheet_count)
copy_and_filter_acordding_to_one_paramter(E_value="5",      Q_value="25",   G_value=None ,   sheet_count=sheet_count)




