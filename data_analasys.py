import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart import LineChart, BarChart, Reference , ScatterChart, Series
from scipy.special import erf
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout

#####################################################################
directory_names = ["Epsilon_3.116_Q_25_G_1.3",
                   "Epsilon_3.116_Q_25_G_1.4",
                   "Epsilon_3.116_Q_25_G_1.5",
                   "Epsilon_3.116_Q_25_G_1.6",
                   "Epsilon_3.116_Q_25_G_1.35",
                   "Epsilon_3.116_Q_25_G_1.375",
                   "Epsilon_3.116_Q_50_G_1.2",
                   "Epsilon_3.116_Q_62.5_G_1.2",
                   "Epsilon_3.116_Q_87.5_G_1.2",
                   #"Epsilon_3.116_Q_75_G_1.2",
                   "Epsilon_3.116_Q_100_G_1.2",
                   "Epsilon_5_Q_50_G_1.2",
                   "Epsilon_3.116_Q_50_G_1.2_bc",
                   "Epsilon_2.077_Q_50_G_1.2",
                   "Epsilon_4.155_Q_50_G_1.2",
                   "Epsilon_5_Q_50_G_1.3",
                   "Epsilon_3.116_Q_50_G_1.3",
                   "Epsilon_4.155_Q_50_G_1.3",
                   "Epsilon_3.116_Q_50_G_1.2_sharpe",
                   "Epsilon_7_Q_50_G_1.2",
                   "Epsilon_6_Q_50_G_1.3",
                   "Epsilon_6_Q_50_G_1.2",
                   "Epsilon_2.077_Q_50_G_1.3",
                   "Epsilon_3.116_Q_81.25_G_1.2",
                   "Epsilon_3.116_Q_68.75_G_1.2",
                   "Epsilon_3.116_Q_50_G_1.4",
                   "Epsilon_4.155_Q_50_G_1.4",
                   "Epsilon_5_Q_50_G_1.4",
                   "Epsilon_6_Q_50_G_1.4",
                   "Epsilon_2.077_Q_50_G_1.4",
                   "Epsilon_3.116_Q_25_G_1.45",
                   #"Epsilon_3.116_Q_25_G_1.55",
                   "Epsilon_3.5_Q_50_G_1.3",
                   "Epsilon_3.5_Q_50_G_1.4",
                   "Epsilon_2.5_Q_50_G_1.4",
                   "Epsilon_2.5_Q_50_G_1.3",
                   "Epsilon_3.116_Q_75_G_1.2_take_2",
                   "Epsilon_3.116_Q_125_G_1.2",
                   "Epsilon_3.116_Q_150_G_1.2",
                   "Epsilon_3.75_Q_50_G_1.3",
                   "Epsilon_2.75_Q_50_G_1.3",
                   "Epsilon_2.75_Q_50_G_1.4",
                   "Epsilon_3.75_Q_50_G_1.4",
                   "Epsilon_2.077_Q_50_G_1.5",
                   "Epsilon_2.5_Q_50_G_1.5",
                   "Epsilon_2.75_Q_50_G_1.5",
                   "Epsilon_3.116_Q_50_G_1.5",
                   "Epsilon_3.5_Q_50_G_1.5",
                   "Epsilon_3.75_Q_50_G_1.5",
                   "Epsilon_4.155_Q_50_G_1.5",
                   "Epsilon_5_Q_50_G_1.5",
                   "Epsilon_3.116_Q_25_G_1.55_take_2",
                   "Epsilon_2.5_Q_25_G_1.4",
                   "Epsilon_2.75_Q_25_G_1.4",
                   "Epsilon_3.116_Q_25_G_1.4_take_2",
                   "Epsilon_3.5_Q_25_G_1.4",
                   "Epsilon_3.75_Q_25_G_1.4",
                   "Epsilon_4.155_Q_25_G_1.4",
                   "Epsilon_5_Q_25_G_1.4",
                   "Epsilon_2.077_Q_25_G_1.4",
                   "Epsilon_5_Q_25_G_1.2",
                   "Epsilon_5_Q_25_G_1.3",
                   "Epsilon_5_Q_25_G_1.35",
                   "Epsilon_5_Q_25_G_1.45",
                   "Epsilon_5_Q_25_G_1.5",
                   "Epsilon_5_Q_25_G_1.55"
                   ]
######################################################################

def process_data(file_path, filter_value, bins_count):
    """Process data: filter, calculate histogram, PDF, CDF, and statistics."""
    df = pd.read_csv(file_path)
    original_data = df.iloc[:, 0]  # Assume the first column is the data

    if filter_value is not None:
        filtered_data = original_data[original_data <= filter_value].dropna().reset_index(drop=True)
    else:
        filtered_data = original_data.dropna().reset_index(drop=True)  # No filtering

    if bins_count>0:
        bins = np.arange(1, bins_count + 1)
    else:
        bins = np.arange(bins_count, -(bins_count - 1))
    
    hist, _ = np.histogram(filtered_data, bins=bins)

    if filtered_data.empty:
        avg, std = None, None
        pdf = np.zeros_like(bins, dtype=float)
        cdf = np.zeros_like(bins, dtype=float)
    else:
        avg = filtered_data.mean()
        std = filtered_data.std(ddof=0)
        if std > 0:
            pdf = (1 / (std * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((bins - avg) / std) ** 2)
            cdf = (1 + erf((bins - avg) / (std * np.sqrt(2)))) / 2
        else:
            pdf = np.zeros_like(bins, dtype=float)
            cdf = np.zeros_like(bins, dtype=float)

    return original_data, filtered_data, bins, hist, pdf, cdf, avg, std

def write_to_worksheet(ws, start_col, data_dict, start_row=2):
    """Write processed data to the worksheet starting at a specific column."""
    for col_num, (col_name, col_data) in enumerate(data_dict.items(), start=start_col):
        ws.cell(row=1, column=col_num).value = col_name
        for row_num, value in enumerate(col_data, start=start_row):
            ws.cell(row=row_num, column=col_num).value = value

def create_combination_chart_1(ws, title, cdf_col, position, bins_count):
    """Create a combination chart with a histogram (bar chart) and scatter plots (PDF and CDF)."""
    bins_col=cdf_col-3
    hist_col=cdf_col-2
    pdf_col=cdf_col-1

    # Bar chart for the histogram
    bar_chart = BarChart()
    bar_chart.title = title
    bar_chart.style = 13
    bar_chart.x_axis.title = "Bins"
    bar_chart.y_axis.title = "Frequency"
    bar_chart.width = 20
    bar_chart.height = 15
    # Hide X-axis for the scatter chart
    #bar_chart.x_axis.delete = True  # Removes the X-axis from the scatter chart

    # Data range for the histogram
    if bins_count > 0:
        hist_data = Reference(ws, min_col=hist_col, min_row=1, max_row=bins_count + 1)
        bins_data = Reference(ws, min_col=bins_col, min_row=2, max_row=bins_count + 1)
        
    else:
        hist_data = Reference(ws, min_col=hist_col, min_row=1, max_row=(-(bins_count*2)) + 1)
        bins_data = Reference(ws, min_col=bins_col, min_row=2, max_row=(-(bins_count*2)) + 1)
        
    
    bar_chart.add_data(hist_data, titles_from_data=True)
    bar_chart.set_categories(bins_data)
    #bar_chart.y_axis.majorGridlines = None

    # Scatter chart for PDF and CDF
    Line_chart_1 = LineChart()
    Line_chart_1.y_axis.title = "Probability"
    Line_chart_1.y_axis.axId = 200  # Secondary Y-axis
    Line_chart_1.y_axis.crosses = "max"  # Align with BarChart

    # Data range for PDF and CDF
    if bins_count > 0:
        pdf_data = Reference(ws, min_col=pdf_col, min_row=1, max_row=bins_count + 1)
        cdf_data = Reference(ws, min_col=cdf_col, min_row=1, max_row=bins_count + 1)
    else:
        pdf_data = Reference(ws, min_col=pdf_col, min_row=1, max_row=(-(bins_count*2)) + 1)
        cdf_data = Reference(ws, min_col=cdf_col, min_row=1, max_row=(-(bins_count*2)) + 1)

    # PDF series
    Line_chart_1.add_data(pdf_data, titles_from_data=True)
    pdf_series = Line_chart_1.series[0]
    pdf_series.graphicalProperties.line.solidFill = "FF0000"  # Red
    pdf_series.graphicalProperties.line.width = 12700 # 1 point (thinner line)

    Line_chart_2 = LineChart()
    Line_chart_2.y_axis.axId = 200  # Secondary Y-axis
    Line_chart_2.y_axis.crosses = "max"  # Align with BarChart
    # CDF series

    Line_chart_2.add_data(cdf_data, titles_from_data=True)
    cdf_series = Line_chart_2.series[0]
    cdf_series.graphicalProperties.line.solidFill = "0000FF"  # Blue
    cdf_series.graphicalProperties.line.width = 12700  # 1 point (thinner line)

    # Combine the charts
    bar_chart += Line_chart_1  # Overlay scatter_chart onto bar_chart
    bar_chart += Line_chart_2

    # Add the combined chart to the worksheet
    ws.add_chart(bar_chart, position)

def process_data_cell_dimensions(file_path, bins_count):
    """Process data without filtering: calculate a ratio, histogram, PDF, CDF, and statistics."""
    df = pd.read_csv(file_path)

    # Ensure there are at least two columns
    if df.shape[1] < 2:
        raise ValueError("The CSV file must contain at least two columns.")

    # Read the first two columns
    col1 = df.iloc[:, 0]
    col2 = df.iloc[:, 1]

    # Avoid division by zero
    if (col2 == 0).any():
        raise ValueError("Division by zero encountered in the second column.")

    # Create a new column as col1 / col2
    ratio_data = col1 / col2

    # Drop NaN values
    ratio_data = ratio_data.dropna().reset_index(drop=True)

    # Define bins for histogram
    bins = np.arange(0, bins_count+0.05, 0.05)  # Bins from 0 to 3 with a step of 0.05

    hist, _ = np.histogram(ratio_data, bins=bins)

    if ratio_data.empty:
        avg, std = None, None
        pdf = np.zeros_like(bins, dtype=float)
        cdf = np.zeros_like(bins, dtype=float)
    else:
        avg = ratio_data.mean()
        std = ratio_data.std(ddof=0)
        if std > 0:
            pdf = (1 / (std * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((bins - avg) / std) ** 2)
            cdf = (1 + erf((bins - avg) / (std * np.sqrt(2)))) / 2
        else:
            pdf = np.zeros_like(bins, dtype=float)
            cdf = np.zeros_like(bins, dtype=float)

     
    # Return results
    return col1 , col2, ratio_data , bins , hist , pdf , cdf ,avg , std

def process_data_to_excel(directory_names, filter_value_w, filter_value_l, output_excel, bins_w, bins_l, bins_t ,bins_cell_dimensions):
    wb = Workbook()
    for idx, dir_name in enumerate(directory_names):
        if idx == 0:
            ws = wb.active
            ws.title = f"{dir_name}"
        else:
            ws = wb.create_sheet(title=f"{dir_name}")

        # Process width data:

        #special filter cases- defualt is filter_value_w=40, filter_value_l=90
        if dir_name=="Epsilon_2.077_Q_50_G_1.2":
            filter_value_w=bins_w=80
            filter_value_l=bins_l=125
        elif dir_name=="Epsilon_3.116_Q_25_G_1.6" or dir_name=="Epsilon_5_Q_50_G_1.3":
            filter_value_w=bins_w=55
        elif dir_name=="Epsilon_3.116_Q_100_G_1.2":
            filter_value_l=bins_l=65
        elif dir_name=="Epsilon_3.116_Q_25_G_1.5" or dir_name=="Epsilon_5_Q_50_G_1.2" or dir_name=="Epsilon_3.116_Q_50_G_1.5" or dir_name=="Epsilon_2.75_Q_50_G_1.5" or dir_name=="Epsilon_2.077_Q_50_G_1.5" or dir_name=="Epsilon_3.116_Q_25_G_1.55_take_2" or dir_name=="Epsilon_3.5_Q_50_G_1.5" or dir_name=="Epsilon_4.155_Q_25_G_1.4" or dir_name=="Epsilon_2.077_Q_25_G_1.4":
            filter_value_w=bins_w=45
        elif dir_name=="Epsilon_3.116_Q_50_G_1.2_sharpe":
            filter_value_w=bins_w=35
            filter_value_l=bins_l=60
        elif dir_name=="Epsilon_7_Q_50_G_1.2":
            filter_value_w=bins_w=70
        elif dir_name=="Epsilon_3.116_Q_150_G_1.2" or dir_name=="Epsilon_4.155_Q_50_G_1.3" or dir_name=="Epsilon_5_Q_50_G_1.3" or dir_name=="Epsilon_5_Q_50_G_1.4" or dir_name=="Epsilon_5_Q_25_G_1.4":
            filter_value_l=bins_l=70
        elif dir_name=="Epsilon_2.077_Q_50_G_1.5" or dir_name=="Epsilon_3.75_Q_50_G_1.5" or dir_name=="Epsilon_4.155_Q_50_G_1.5" or dir_name=="Epsilon_5_Q_50_G_1.5" or dir_name=="Epsilon_5_Q_25_G_1.5":
            filter_value_w=bins_w=60
        elif dir_name=="Epsilon_3.5_Q_50_G_1.3":
            filter_value_l=bins_l=55
        elif dir_name=="Epsilon_3.116_Q_25_G_1.3" or dir_name=="Epsilon_3.116_Q_62.5_G_1.2" or dir_name=="Epsilon_3.116_Q_68.75_G_1.2" or dir_name=="Epsilon_3.116_Q_75_G_1.2" or dir_name=="Epsilon_3.116_Q_75_G_1.2_take_2" or dir_name=="Epsilon_3.116_Q_81.25_G_1.2" or dir_name=="Epsilon_3.116_Q_50_G_1.3" or dir_name=="Epsilon_3.116_Q_50_G_1.2" or  dir_name=="Epsilon_3.116_Q_87.5_G_1.2" or dir_name=="Epsilon_5_Q_50_G_1.2" or dir_name=="Epsilon_6_Q_50_G_1.2" or dir_name=="Epsilon_6_Q_50_G_1.4" or dir_name=="Epsilon_3.116_Q_50_G_1.2_bc" or dir_name=="Epsilon_5_Q_25_G_1.3" or dir_name=="Epsilon_5_Q_25_G_1.35" :
            filter_value_l=bins_l=60
        elif dir_name=="Epsilon_2.5_Q_50_G_1.5" or dir_name=="Epsilon_3.75_Q_50_G_1.3" or dir_name=="Epsilon_3.75_Q_50_G_1.4" or dir_name=="Epsilon_3.116_Q_25_G_1.5" or dir_name=="Epsilon_4.155_Q_50_G_1.4" or dir_name=="Epsilon_5_Q_50_G_1.4" or dir_name=="Epsilon_6_Q_50_G_1.3" or dir_name=="Epsilon_5_Q_25_G_1.4" or dir_name=="Epsilon_5_Q_25_G_1.3" or dir_name=="Epsilon_5_Q_25_G_1.35" or dir_name=="Epsilon_5_Q_25_G_1.45" :
            filter_value_w=bins_w=50
        elif dir_name=="Epsilon_5_Q_25_G_1.2":
            filter_value_l=bins_l=50
        elif dir_name=="Epsilon_5_Q_25_G_1.55":
            filter_value_w=bins_w=50
            filter_value_l=bins_l=70
            
        file_path_w = os.path.join(dir_name, "hist_euclidean_width.csv")
        original, filtered, bins, hist, pdf, cdf, avg, std = process_data(file_path_w, filter_value_w, bins_w)
        data_dict_w = {
            "Width": original.tolist(),
            "Filtered": filtered.tolist(),
            "Bins": bins.tolist(),
            "Hist": hist.tolist(),
            "PDF": pdf.tolist(),
            "CDF": cdf.tolist(),
            "Avg/Std": [avg, std]
        }
        write_to_worksheet(ws, start_col=1, data_dict=data_dict_w)
        create_combination_chart_1(ws, f"PDF and CDF Width - {dir_name}", cdf_col=6, position="H2", bins_count=bins_w)

        # Process length data
        file_path_l = os.path.join(dir_name, "hist_euclidean_length.csv")
        original, filtered, bins, hist, pdf, cdf, avg, std = process_data(file_path_l, filter_value_l, bins_l)
        data_dict_l = {
            "Length": original.tolist(),
            "Filtered": filtered.tolist(),
            "Bins": bins.tolist(),
            "Hist": hist.tolist(),
            "PDF": pdf.tolist(),
            "CDF": cdf.tolist(),
            "Avg/Std": [avg, std]
        }
        write_to_worksheet(ws, start_col=19, data_dict=data_dict_l)
        create_combination_chart_1(ws, f"PDF and CDF Length - {dir_name}", cdf_col=24, position="Z2", bins_count=bins_l)

        # Process theta data
        file_path_t = os.path.join(dir_name, "hist_theta_values.csv")
        original, filtered, bins, hist, pdf, cdf, avg, std = process_data(file_path_t, None, bins_t)
        data_dict_l = {
            "Theta": original.tolist(),
            "Bins": bins.tolist(),
            "Hist": hist.tolist(),
            "PDF": pdf.tolist(),
            "CDF": cdf.tolist(),
            "Avg/Std": [avg, std]
        }
        write_to_worksheet(ws, start_col=37, data_dict=data_dict_l)
        create_combination_chart_1(ws, f"PDF and CDF Theta - {dir_name}", cdf_col=41, position="AR2", bins_count=bins_t)


        # Process cell dimensions data
        file_path_cell_dimensions = os.path.join(dir_name, "cell_dimensions.csv")
        col1, col2, ratio, bins, hist, pdf, cdf, avg, std = process_data_cell_dimensions(file_path_cell_dimensions, bins_cell_dimensions)
        data_dict_cell_dimensions = {
            "Length": col1.tolist(),
            "Width": col2.tolist(),
            "Ratio":ratio.tolist(),
            "Bins": bins.tolist(),
            "Hist": hist.tolist(),
            "PDF": pdf.tolist(),
            "CDF": cdf.tolist(),
            "Avg/Std": [avg, std]
        }

        bins = (bins_cell_dimensions/0.05)  #this is to correctlly get the amount of bins for the graphs. its the bins devided by the step
        write_to_worksheet(ws, start_col=55, data_dict=data_dict_cell_dimensions)
        create_combination_chart_1(ws, f"PDF and CDF L/W - {dir_name}", cdf_col=61, position="BK2", bins_count=bins)

        #reset the defualt values of  filters
        filter_value_w=40
        filter_value_l=90
        bins_w=40
        bins_l=90


    wb.save(output_excel)
    print(f"Data processed and saved to {output_excel}")


# Example usage
process_data_to_excel(directory_names, filter_value_w=40, filter_value_l=90, output_excel="data_analasys.xlsx", bins_w=50, bins_l=90, bins_t=-30 ,bins_cell_dimensions= 3)
